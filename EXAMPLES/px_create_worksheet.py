#!/usr/bin/env python
import openpyxl as px

fruits = ["pomegranate", "cherry", "apricot", "date", "apple", "lemon", "kiwi",
          "orange", "lime", "watermelon", "guava", "papaya", "fig", "pear", "banana",
          "tamarind", "persimmon", "elderberry", "peach", "blueberry", "lychee",
          "grape"]

wb = px.Workbook()

ws = wb.active

ws.title = 'fruits'

for i, fruit in enumerate(fruits, 1):
    column_1 = ws.cell(row=i, column=1)
    column_1.value = fruit
    column_1.font=px.styles.Font(color='00FF0000', name="Comic Sans", size=18)

    column_2 = ws.cell(row=i, column=2)
    column_2.value = len(fruit)

ws2 = wb.create_sheet("other")
ws2.cell(row=1, column=1).value = "Testing..."

wb.save('fruits.xlsx')
